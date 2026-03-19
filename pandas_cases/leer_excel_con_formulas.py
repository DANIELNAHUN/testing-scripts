import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path

FILE_PATH = Path(__file__).parent.parent / "files" / "source" / "Data_Insatisfaccion_Enero.xlsx"

def cargar_umbrales(wb):
    ws = wb["Umbrales"]

    # Pesos: columna I = nombre, columna J = peso (filas 2 a 11)
    pesos = {}
    for i in range(2, 12):
        nombre = ws[f"I{i}"].value
        peso   = ws[f"J{i}"].value
        if nombre and peso is not None:
            pesos[nombre] = peso

    # Rangos cualitativos: J16:K20, orden descendente para BUSCARV
    rangos = []
    for i in range(16, 21):
        umbral  = ws[f"J{i}"].value
        etiqueta = ws[f"K{i}"].value
        if umbral is not None:
            rangos.append((umbral, etiqueta))
    rangos.sort(key=lambda x: x[0], reverse=True)  # mayor a menor

    return pesos, rangos

def cargar_mapeo_score(wb):
    """
    Lee la fórmula de Score en N2 para extraer dinámicamente el mapeo
    columna_data -> fila_umbrales, y devuelve lista de (header_data, peso).
    La fórmula tiene la forma: (C2*J4 + D2*J7 + E2*J10 + ...) / SUM(J2:J11) * 100
    """
    import re
    ws_data = wb["Data"]
    ws_umb  = wb["Umbrales"]

    formula = ws_data["N2"].value  # fórmula de Score en primera fila de datos
    # Extraer pares (col_data, fila_j): ej. C2*$Umbrales.$J$4 -> (C, 4)
    patron = re.findall(r'\$?([A-J])\d+\*[^$]*\$J\$(\d+)', formula)

    mapeo = []
    for col_letra, fila_j in patron:
        header = ws_data[f"{col_letra}1"].value
        peso   = ws_umb[f"J{fila_j}"].value
        mapeo.append((header, peso))
    return mapeo

def es_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def detectar_columnas_formula(wb):
    """
    Revisa la fila 2 de las columnas Score (N), Score Cualitativo (O) y Causas (P)
    y devuelve cuáles son fórmula y cuáles tienen valor fijo.
    """
    ws = wb["Data"]
    return {
        "Score":             es_formula(ws["N2"]),
        "Score Cualitativo": es_formula(ws["O2"]),
        "Causas":            es_formula(ws["P2"]),
    }

def calcular_score(row, mapeo_score, suma_pesos):
    total = sum(row[header] * peso for header, peso in mapeo_score)
    return (total / suma_pesos) * 100

def calcular_cualitativo(score, rangos):
    if pd.isna(score):
        return None
    for umbral, etiqueta in rangos:
        if score >= umbral:
            return etiqueta
    return None

def calcular_causas(row, mapeo_score):
    cols_score = [header for header, _ in mapeo_score]
    return ", ".join(col for col in cols_score if row[col] > 0) + ","

# --- Main ---
wb = openpyxl.load_workbook(FILE_PATH, data_only=False)

pesos, rangos  = cargar_umbrales(wb)
mapeo_score    = cargar_mapeo_score(wb)
suma_pesos     = sum(pesos.values())
es_formula_col = detectar_columnas_formula(wb)

df = pd.read_excel(FILE_PATH, sheet_name="Data")
df = df.loc[:, ~df.columns.str.startswith("Unnamed")]

# Score
if es_formula_col["Score"]:
    df["Score"] = df.apply(lambda r: calcular_score(r, mapeo_score, suma_pesos), axis=1)
else:
    df["Score"] = pd.read_excel(FILE_PATH, sheet_name="Data", usecols=["Score"])["Score"]

# Score Cualitativo
if es_formula_col["Score Cualitativo"]:
    df["Score Cualitativo"] = df["Score"].apply(lambda s: calcular_cualitativo(s, rangos))
else:
    df["Score Cualitativo"] = pd.read_excel(FILE_PATH, sheet_name="Data", usecols=["Score Cualitativo"])["Score Cualitativo"]

# Causas
if es_formula_col["Causas"]:
    df["Causas"] = df.apply(lambda r: calcular_causas(r, mapeo_score), axis=1)
else:
    df["Causas"] = pd.read_excel(FILE_PATH, sheet_name="Data", usecols=["Causas"])["Causas"]

print(f"Total registros: {len(df)}")
print(f"Score es fórmula:             {es_formula_col['Score']}")
print(f"Score Cualitativo es fórmula: {es_formula_col['Score Cualitativo']}")
print(f"Causas es fórmula:            {es_formula_col['Causas']}\n")

sample_json = df.head(5).to_json(orient="records", force_ascii=False, indent=2, double_precision=12)
print("Primeros 5 registros:")
print(sample_json)
